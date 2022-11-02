from ast import Num
from operator import itemgetter
from random import choices, randrange, sample, shuffle
import random
from unittest import skip
from numpy import datetime_as_string
import numpy
from openpyxl import load_workbook
import pandas as pd
from time import time, gmtime, strftime
from tabulate import tabulate

NUM_OF_PEOPLE = 10000

NUM_OF_DISTRICTS = 46        # 4628 TODO?
TIME = 8                    # in days (TODO: min 200)
MIN_AGE = 12                # TODO: child later
MAX_AGE = 113
# DISTRICT_CAPACITY = 1     # TODO?
SHIPMENT = 7
SHIPMENT_START_DATE = "2021-02-16"
SHIPMENT_AMOUNT = 0
WAIT = 7
CHRONIC_DISEASE_LIST = ["yes", "no"]
PERC_OF_ALL_PEOPLE_WITH_CHRONIC_DIS = 25
VACCINE_LIST = ["pfizer", "moderna", "astrazeneca", "sputnik", "sinopharm"]
VACCINE_LIST_D = ["yes", "yes", "yes", "no", "no"]
VACCINE_LIST_M = [1, 0.12, 0.4, 0.16, 0.32]

people_list = []
vacc_list = []
district_list = []


def to_xlsx(my_list, name: str):
    df = pd.DataFrame.from_dict(my_list)
    df.to_excel(f"{name}.xlsx")


def vacc_shipment(name: str, SSD):
    vacc_shipment_excel = pd.read_excel('data/vacc_shipment.xlsx')
    data = pd.DataFrame(vacc_shipment_excel, columns=[
                        'id', 'date', 'pfizer', 'moderna', 'astrazeneca', 'sputnik', 'sinopharm'])
    idx = data.index[data['date'] == SSD]

    # kerekítés így oké?
    SHIPMENT_AMOUNT = round(data[name].values[idx[0]] / 1000)
    SHIPMENT_COUNTER = idx[0]
    SHIPMENT_COUNTER += SHIPMENT
    _SSD = data["date"].values[SHIPMENT_COUNTER]
    SHIPMENT_START_DATE = datetime_as_string(_SSD, unit='D')

    return SHIPMENT_AMOUNT


def generate_vaccines():
    for x in range(len(VACCINE_LIST)):
        dict = {
            "name": VACCINE_LIST[x],
            # TODO: krónikus beteg utána néz és beállít
            "can_get_chronic_disease": "".join(VACCINE_LIST_D[x]),
            "pref_age": [MIN_AGE+4, MAX_AGE] if not VACCINE_LIST[x] == "pfizer" else [MIN_AGE+6, MAX_AGE],
            "num_of_vacc": vacc_shipment(VACCINE_LIST[x], SHIPMENT_START_DATE)
        }

        vacc_list.append(dict)

    to_xlsx(vacc_list, "original_vaccines")

    return vacc_list


def generate_people(num_of_people: int):
    if num_of_people == 0:
        print("The number of people should be greater than 0!")
        return

    for x in range(num_of_people):
        pref_num = randrange(0, len(VACCINE_LIST)+1)
        pref_pairs_list = []
        pref_vacc = sample(VACCINE_LIST, pref_num)

        for i in range(pref_num):
            # TODO: kellene valami arány?
            pref_pairs_list.append((pref_vacc[i], randrange(1, 100)))

        dict = {
            "id": x + 1,
            "age": randrange(MIN_AGE, MAX_AGE),
            "chronic_disease": "".join(choices(CHRONIC_DISEASE_LIST, weights=(PERC_OF_ALL_PEOPLE_WITH_CHRONIC_DIS, 100-PERC_OF_ALL_PEOPLE_WITH_CHRONIC_DIS))),
            "pref_list": list(sorted(pref_pairs_list, key=lambda x: x[1], reverse=True)),
            "district": None
        }

        people_list.append(dict)

    to_xlsx(people_list, "original_people")

    return people_list


def splitter(p_list, size):
    return list(p_list[i::size] for i in range(size))


def find_by_id(id):
    people = next((i for i in people_list if i['id'] == id), None)
    return people


def assign_people_to_districts(districts):
    people_id_list = []

    for i in people_list:
        people_id_list.append(i["id"])

    for x in range(districts):
        dict = {
            "id": x + 1,
            "id_list": splitter(people_id_list, districts)[x]
        }

        district_list.append(dict)

    for d in district_list:
        distr = d["id"]
        list = d["id_list"]

        for id in list:
            p = find_by_id(id)
            people_list[p["id"]-1]["district"] = distr

    to_xlsx(district_list, "districts")
    to_xlsx(people_list, "people_with_districts")

    return district_list


# TODO: minden vakcinából választunk random valamennyit - hogy máshogy?
def choose_random_vacc(vacc_list_m, a_v_list=[]):
    for x in range(len(VACCINE_LIST)):
        # magic number, 50 jó átlagnak tűnt... ajánlat szerint átír
        a = int(10*vacc_list_m[x])
        if vacc_list[x]["num_of_vacc"] >= a:
            while a != 0:
                dict = {
                    "name": vacc_list[x]["name"],
                    "can_get_chronic_disease": vacc_list[x]["can_get_chronic_disease"],
                    "pref_age": vacc_list[x]["pref_age"]
                }
                a_v_list.append(dict)
                a -= 1
        elif vacc_list[x]["num_of_vacc"] < a:
            a = vacc_list[x]["num_of_vacc"]
            while a != 0:
                dict = {
                    "name": vacc_list[x]["name"],
                    "can_get_chronic_disease": vacc_list[x]["can_get_chronic_disease"],
                    "pref_age": vacc_list[x]["pref_age"]
                }
                a_v_list.append(dict)
                a -= 1

    return shuffle(a_v_list)

# TODO: vacc_algorithm csak azt ajánlanák neki amit a legjobban preferál


def vacc_algorithm(days: int):
    d = 1
    temp = []
    available_vaccine_list = []
    not_vacc_people = []
    vacc_people = []

    # TODO: talán kellene olyan tábla hogy ki mivel lett pontosan oltva, mit utasított vissza stb.
    vaccinated_with = []
    for vw in range(len(VACCINE_LIST)):
        dict = {
            "name": VACCINE_LIST[vw],
            "amount": 0
        }
        vaccinated_with.append(dict)

    rejected = []
    for r in range(len(VACCINE_LIST)):
        dict = {
            "name": VACCINE_LIST[r],
            "amount": 0
        }
        rejected.append(dict)

    while d <= days:
        for idx in not_vacc_people:
            idx[1] += 1
        if d % WAIT == 0:
            for n in not_vacc_people:
                if n[1] == WAIT:
                    print("waited")
                    print(n[0])
                    people_list.append(n[0])
                    del not_vacc_people[n]

        for i in range(len(district_list)):
            temp.clear()
            available_vaccine_list.clear()
            people_by_distr = []

            for p in people_list:
                if p["district"] == i+1 and p["pref_list"]:
                    people_by_distr.append(p)

            #temp = sorted(people_by_distr, key=itemgetter('age'), reverse=True)
            temp = sorted(people_by_distr, key=lambda k: (
                k['age'], k['chronic_disease']), reverse=True)
            to_xlsx(temp, "temp")

            choose_random_vacc(VACCINE_LIST_M, available_vaccine_list)
            to_xlsx(available_vaccine_list, "available_vaccine_list")

            while len(available_vaccine_list) > 0:
                p_to_vacc = temp[:1]
                pref_list = []
                pref_list += p_to_vacc[0]["pref_list"]
                age = p_to_vacc[0]["age"]
                vacc = ""
                perc = 0

                for pref in range(len(pref_list)):
                    if pref_list[pref][0] == available_vaccine_list[0]["name"]:
                        in_pref_list = True
                        vacc += available_vaccine_list[0]["name"]
                        perc += pref_list[pref][1]
                        break
                    else:
                        in_pref_list = False

                if in_pref_list == True and (p_to_vacc[0]["chronic_disease"] == available_vaccine_list[0]["can_get_chronic_disease"] or p_to_vacc[0]["chronic_disease"] == "no") and (True if (vacc == "pfizer" and age >= MIN_AGE+4) or (vacc != "pfizer" and age >= MIN_AGE+6) else False):

                    decision = "".join(
                        choices(CHRONIC_DISEASE_LIST, weights=(perc, 100-perc)))

                    if decision == "yes":
                        for vaccinate in vaccinated_with:
                            if vaccinate["name"] == available_vaccine_list[0]["name"]:
                                vaccinate["amount"] += 1

                        for w in range(len(vacc_list)):
                            if available_vaccine_list[0]["name"] == vacc_list[w]["name"]:
                                num = int(vacc_list[w]["num_of_vacc"])
                                vacc_list[w]["num_of_vacc"] = num-1

                        available_vaccine_list.pop(0)

                        for q in range(len(people_list)):
                            if people_list[q]["id"] == p_to_vacc[0]["id"]:
                                del people_list[q]
                                break

                        vacc_people.append(p_to_vacc[0])
                        to_xlsx(vacc_people, "vacc_people")

                        temp.pop(0)
                    else:
                        not_vacc_people.append([p_to_vacc[0], 0])
                        to_xlsx(not_vacc_people, "not_vacc_people")

                        for reject in rejected:
                            if reject["name"] == available_vaccine_list[0]["name"]:
                                reject["amount"] += 1

                        for _q in people_list:
                            if _q["id"] == p_to_vacc[0]["id"]:
                                people_list.remove(_q)

                        available_vaccine_list.pop(0)

                        temp.pop(0)
                else:
                    temp.pop(0)
                    # print("skipped")
                    # print(p_to_vacc[0])
                    # print(pref_list)
                    # print(in_pref_list)

        d += 1

        if d % SHIPMENT == 0:
            for i in vacc_list:
                for x in range(len(VACCINE_LIST)):
                    if VACCINE_LIST[x] == i["name"]:
                        next_amount = vacc_shipment(
                            i["name"], SHIPMENT_START_DATE)
                        if SHIPMENT_AMOUNT == next_amount:
                            break
                        elif SHIPMENT_AMOUNT < next_amount:
                            i["num_of_vacc"] += next_amount-SHIPMENT_AMOUNT
            print("SHIPMENT")

    sorted_people_list = sorted(
        people_list, key=itemgetter('id'), reverse=False)
    to_xlsx(vacc_list, "vaccines_end_of_day")
    to_xlsx(sorted_people_list, "people_end_of_day")

    to_xlsx(rejected, "vacc_rejected")
    to_xlsx(vaccinated_with, "vacc_with")

    header_rejected = rejected[0].keys()
    rows_rejected = [x.values() for x in rejected]
    print("\nRejected vaccines by type:\n",
          tabulate(rows_rejected, header_rejected))

    header_accepted = vaccinated_with[0].keys()
    rows_accepted = [x.values() for x in vaccinated_with]
    print("\nAccepted vaccines by type:\n",
          tabulate(rows_accepted, header_accepted))

    wrkbk_not_vacc_p = load_workbook("not_vacc_people.xlsx")
    sh_not_vacc_p = wrkbk_not_vacc_p.active
    print("\nSum of not vaccinated people: " + str(sh_not_vacc_p.max_row-1))

    wrkbk_vacc_p = load_workbook("vacc_people.xlsx")
    sh_vacc_p = wrkbk_vacc_p.active
    print("Sum of vaccinated people: " + str(sh_vacc_p.max_row-1) + "\n")

    return


if __name__ == "__main__":
    st = time()

    generate_people(NUM_OF_PEOPLE)
    generate_vaccines()
    assign_people_to_districts(NUM_OF_DISTRICTS)

    vacc_algorithm(TIME)

    elapsed_time = time() - st
    print('Execution time:', strftime("%H:%M:%S", gmtime(elapsed_time)), "\n")
